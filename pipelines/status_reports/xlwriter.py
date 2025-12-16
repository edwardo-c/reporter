from openpyxl import load_workbook
from pathlib import Path

SHEET_NAME = "Summary"

def generate_reports(
        report_maps: dict[str, dict[str, float]],
        templates: dict[bool, str],
        out_dir: Path,
        ):
    """
    Iterate over report map to output finished reports

    args:
        report_maps, expected structure: -    
            [
                {
                    'meta': {
                        'out_file_name': base name of file to be created, 
                        'signed': bool value, used to determine which template to use, 
                        'acct_num': 'abc123' - not used, only for reference
                        },
                    'value_map: {A1: 123, B2: 456}
                }, 
            ]

        templates: {
                        true: 'full/signed/template/path.xlsx'
                        false: 'full/UNSIGNED/template/path.xlsx'
        } 

        out_dir: directory to save finished reports in
    """

    for cust in report_maps:
    
        meta = cust["meta"]
        report_map = cust["value_map"]

        out_path = out_dir / meta["out_file_name"]

        template_path = templates[True] if bool(meta["signed"]) else templates[False]
        
        _fill_report_template(template_path, out_path, report_map)

def _fill_report_template(
        path_in: str, path_out: Path,
        report_map: dict[str, float]
    ):    
    wb = load_workbook(path_in)
    ws = wb[SHEET_NAME]
    for range, val in report_map.items():
        ws[range].value = val
    wb.save(path_out)
    wb.close()

