import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

@pytest.fixture
def two_sheet_xlsx(tmp_path) -> str:
    """
    fixture of xlsx with two sheets
    each sheet has three columns of two rows: 9 values total including headers
    sheet_1 starts at row 1
    sheet_2 starts at row 10
    """
    out_path = str(tmp_path / "t.xlsx")

    wb = Workbook()
    
    def _write_to_excel(
            data: list[list[str | float | int | None]],
            ws: Worksheet, start_row: int):
        """write a list of lists to existing sheet, specify start row"""
        for r, row in enumerate(data, start=start_row):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)

    wb.remove(wb.active)

    wb.create_sheet("sheet_1")
    wb.create_sheet("sheet_2")

    sheet_1 = wb["sheet_1"]
    sheet_2 = wb["sheet_2"]

    data_1 = [
        ['Inventory CD', 'Classification(category)', 'Amount'],
        ['a123', 'electronic', 100],
        ['b456', 'furniture', 250]            
    ]

    data_2 = [
        ['Part Number', 'sales category', 'total'],
        ['d456', 'electronic', 35],
        ['e456', 'lighting', 180]            
    ]

    _write_to_excel(data_1, sheet_1, 1)
    _write_to_excel(data_2, sheet_2, 10)


    wb.save(out_path)

    return out_path

def test_single_file(two_sheet_xlsx):
    """
    test: 
    - reading single file
    - renaming of columns
    - stacking of dataframes
    """
    ...