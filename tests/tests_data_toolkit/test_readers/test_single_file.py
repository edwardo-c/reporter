from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import pytest

from data_toolkit.readers.single_file import read_data

def test_single_file_reader(
        cfg: dict[str, Any], 
        expected_columns: set[str], 
        expected_results: pd.DataFrame
    ):
    """
    test: 
    - reading single file with two sheets starting at different rows
    - renaming of columns
    - stacking of dataframes
    """
    df: pd.DataFrame = read_data(cfg=cfg)

    assert expected_columns <= set(df.columns)

    pd.testing.assert_frame_equal(df, expected_results, check_like=True)

@pytest.fixture
def expected_columns() -> set[str]:
    """Return set of expected columns through cfg fixture"""
    return {'part_number', 'category', 'amount'}

@pytest.fixture
def expected_results() -> pd.DataFrame:
    """Expected dataframe post stacking of through cfg"""
    return pd.DataFrame(
        {
            'part_number': ['a123', 'b456', 'd456', 'e456'], 
            'category': ['electronic', 'furniture', 'electronic', 'lighting'], 
            'amount': [100, 250, 35, 180]
        }
    )

@pytest.fixture
def cfg(two_sheet_xlsx) -> dict[str, list[dict[str, str | list | dict]]]:
    return {
        'file_path': two_sheet_xlsx,
        'params': [
            {
                'sheet_name': 'sheet_1',
                'header': 0,
                'usecols': ['Inventory CD', 'Classification(category)', 'Amount'],
                'rename_map': {
                    'Inventory CD': 'part_number',
                    'Classification(category)': 'category',
                    'Amount': 'amount'
                }
            },
            {
                'sheet_name': 'sheet_2',
                'header': 9,
                'usecols': ['Part Number', 'sales category', 'total'],
                'rename_map': {
                    'Part Number': 'part_number',
                    'sales category': 'category',
                    'total': 'amount'
                }
            }      
        ]
    }

@pytest.fixture
def two_sheet_xlsx(tmp_path) -> str:
    """
    return path of temporary .xlsx file with two sheets
    each sheet has three columns of two rows: 9 values total
    sheet_1 starts at row 1
    sheet_2 starts at row 10
    """

    out_path = str(tmp_path / "t.xlsx")
    wb = Workbook()
    
    def _write_to_excel(
            data: list[list[str | float | int | None]],
            ws: Worksheet, start_row: int):
        """write a list of lists to existing sheet, specify start row"""
        for r, row in enumerate(data, start=start_row):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)

    wb.remove(wb.active)

    wb.create_sheet("sheet_1")
    wb.create_sheet("sheet_2")

    sheet_1 = wb["sheet_1"]
    sheet_2 = wb["sheet_2"]

    data_1 = [
        ['Inventory CD', 'Classification(category)', 'Amount'],
        ['a123', 'electronic', 100],
        ['b456', 'furniture', 250]            
    ]

    data_2 = [
        ['Part Number', 'sales category', 'total'],
        ['d456', 'electronic', 35],
        ['e456', 'lighting', 180]            
    ]

    _write_to_excel(data_1, sheet_1, 1)
    _write_to_excel(data_2, sheet_2, 10)


    wb.save(out_path)

    return out_path